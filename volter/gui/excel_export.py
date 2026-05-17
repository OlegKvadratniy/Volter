from __future__ import annotations

from pathlib import Path

from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from PySide6.QtCore import Qt, Signal


class ExcelExportScreen(QWidget):
    go_to_setup = Signal()

    def __init__(self) -> None:
        super().__init__()
        self._products: list[dict] = []
        self._order_data: dict = {}
        self._weights: dict[str, float] = {}
        self._packaging_weight_g: float = 0.0
        self._shipping_rate_usd: float = 0.0
        self._init_ui()

    def _init_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(40, 40, 40, 40)

        title = QLabel("Excel", alignment=Qt.AlignCenter)
        title.setStyleSheet("font-size: 24px; font-weight: bold;")
        layout.addWidget(title)

        self.order_label = QLabel("Заказ не создан")
        self.order_label.setStyleSheet("color: #888; font-style: italic;")
        layout.addWidget(self.order_label)

        rate_layout = QHBoxLayout()
        rate_layout.addWidget(QLabel("Курс юань/$:"))
        self.rate_input = QLineEdit("7.2")
        self.rate_input.setFixedWidth(80)
        self.rate_input.editingFinished.connect(self._refresh_table)
        rate_layout.addWidget(self.rate_input)
        rate_layout.addStretch()
        layout.addLayout(rate_layout)

        layout.addWidget(QLabel("Товары:"))
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        headers = ["№", "Фото", "Название RU", "Цена ¥", "Цена $", "SKU", "Заказчик", "Описание", "Кол-во", "Сумма $", "Ссылка", "Статус"]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setColumnWidth(0, 40)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(3, 70)
        self.table.setColumnWidth(4, 70)
        self.table.setColumnWidth(8, 60)
        self.table.setColumnWidth(11, 70)
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)

        layout.addWidget(QLabel("Вес участников:"))
        self.weight_table = QTableWidget()
        self.weight_table.setColumnCount(7)
        self.weight_table.setHorizontalHeaderLabels(["Участник", "Вес (кг)", "Упаковка (кг)", "Итого вес (кг)", "Доставка ($)", "Товары ($)", "Итого ($)"])
        self.weight_table.setColumnWidth(0, 120)
        self.weight_table.setColumnWidth(1, 80)
        self.weight_table.setColumnWidth(2, 100)
        self.weight_table.setColumnWidth(3, 100)
        self.weight_table.setColumnWidth(4, 100)
        self.weight_table.setColumnWidth(5, 90)
        self.weight_table.setColumnWidth(6, 90)
        self.weight_table.setFixedHeight(200)
        layout.addWidget(self.weight_table)

        btn_layout = QHBoxLayout()

        self.export_products_btn = QPushButton("Экспорт товаров")
        self.export_products_btn.setFixedHeight(40)
        self.export_products_btn.setStyleSheet("font-size: 14px; font-weight: bold; background-color: #2196F3; color: white;")
        self.export_products_btn.clicked.connect(self._export_products)
        btn_layout.addWidget(self.export_products_btn)

        self.export_weights_btn = QPushButton("Экспорт веса")
        self.export_weights_btn.setFixedHeight(40)
        self.export_weights_btn.setStyleSheet("font-size: 14px; font-weight: bold; background-color: #FF9800; color: white;")
        self.export_weights_btn.clicked.connect(self._export_weights)
        btn_layout.addWidget(self.export_weights_btn)

        self.new_order_btn = QPushButton("Новый заказ")
        self.new_order_btn.setFixedHeight(40)
        self.new_order_btn.clicked.connect(lambda: self.go_to_setup.emit())
        btn_layout.addWidget(self.new_order_btn)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)

    def set_order(self, order_data: dict) -> None:
        self._order_data = order_data
        self.order_label.setText(f"Заказ: {order_data['name']}")
        self.order_label.setStyleSheet("color: #333; font-weight: bold;")
        self._refresh_weight_table()

    def set_weights(self, weights: dict) -> None:
        self._weights = weights.get("weights", {})
        self._packaging_weight_g = weights.get("packaging_weight_g", 0.0)
        self._shipping_rate_usd = weights.get("shipping_rate_usd", 0.0)
        self._refresh_weight_table()

    def add_product(self, product_data: dict) -> None:
        self._products.append(product_data)
        self._refresh_table()

    def _get_rate(self) -> float:
        try:
            return float(self.rate_input.text())
        except ValueError:
            return 7.2

    def _refresh_table(self) -> None:
        rate = self._get_rate()
        self.table.setRowCount(len(self._products))
        for i, p in enumerate(self._products):
            self.table.setItem(i, 0, QTableWidgetItem(str(i + 1)))

            photo_item = QTableWidgetItem()
            photo_item.setTextAlignment(Qt.AlignCenter)
            image_path = p.get("image_path", "")
            if image_path and Path(image_path).exists():
                from PySide6.QtGui import QPixmap
                pixmap = QPixmap(image_path).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                photo_item.setData(Qt.DecorationRole, pixmap)
            else:
                photo_item.setText("Нет")
            self.table.setItem(i, 1, photo_item)

            title = p.get("title_ru", "") or p.get("title_zh", "")
            self.table.setItem(i, 2, QTableWidgetItem(title[:60]))

            price_yuan = p["price_yuan"]
            price_usd = round(price_yuan / rate, 2)
            self.table.setItem(i, 3, QTableWidgetItem(str(price_yuan)))
            self.table.setItem(i, 4, QTableWidgetItem(str(price_usd)))

            self.table.setItem(i, 5, QTableWidgetItem(p.get("sku_selected", "")))
            self.table.setItem(i, 6, QTableWidgetItem(p.get("customer", "")))
            self.table.setItem(i, 7, QTableWidgetItem(p.get("description", "")))

            qty = p.get("quantity", 1)
            self.table.setItem(i, 8, QTableWidgetItem(str(qty)))

            total_usd = round(price_usd * qty, 2)
            self.table.setItem(i, 9, QTableWidgetItem(str(total_usd)))

            self.table.setItem(i, 10, QTableWidgetItem(p.get("full_url", "")[:50]))

            status_str = "Есть" if p.get("status", True) else "Нет"
            self.table.setItem(i, 11, QTableWidgetItem(status_str))

            self.table.setRowHeight(i, 90)

    def _refresh_weight_table(self) -> None:
        participants = self._order_data.get("participants", [])
        self.weight_table.setRowCount(len(participants) + 1)

        pkg_share = self._packaging_weight_g / 1000.0 / len(participants) if participants else 0.0
        total_weight = 0.0
        total_delivery = 0.0
        total_goods = 0.0
        total_all = 0.0

        for i, name in enumerate(participants):
            personal = self._weights.get(name, 0.0)
            total_kg = personal + pkg_share
            delivery = total_kg * self._shipping_rate_usd
            goods = self._calc_customer_cost(name)
            grand = goods + delivery
            total_weight += total_kg
            total_delivery += delivery
            total_goods += goods
            total_all += grand

            self.weight_table.setItem(i, 0, QTableWidgetItem(name))
            self.weight_table.setItem(i, 1, QTableWidgetItem(f"{personal:.2f}"))
            self.weight_table.setItem(i, 2, QTableWidgetItem(f"{pkg_share:.2f}"))
            self.weight_table.setItem(i, 3, QTableWidgetItem(f"{total_kg:.2f}"))
            self.weight_table.setItem(i, 4, QTableWidgetItem(f"${delivery:.2f}"))
            self.weight_table.setItem(i, 5, QTableWidgetItem(f"${goods:.2f}"))
            self.weight_table.setItem(i, 6, QTableWidgetItem(f"${grand:.2f}"))

        from PySide6.QtGui import QFont
        bold = QFont()
        bold.setBold(True)

        self.weight_table.setItem(len(participants), 0, QTableWidgetItem("Итого"))
        self.weight_table.setItem(len(participants), 1, QTableWidgetItem(f"{total_weight:.2f}"))
        self.weight_table.setItem(len(participants), 2, QTableWidgetItem(f"{pkg_share * len(participants):.2f}"))
        self.weight_table.setItem(len(participants), 3, QTableWidgetItem(f"{total_weight:.2f}"))
        self.weight_table.setItem(len(participants), 4, QTableWidgetItem(f"${total_delivery:.2f}"))
        self.weight_table.setItem(len(participants), 5, QTableWidgetItem(f"${total_goods:.2f}"))
        self.weight_table.setItem(len(participants), 6, QTableWidgetItem(f"${total_all:.2f}"))
        for col in range(7):
            self.weight_table.item(len(participants), col).setFont(bold)

    def _export_products(self) -> None:
        if not self._products:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт товаров", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import PatternFill, Font, Alignment
            from PIL import Image as PILImage
            import io

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self._order_data.get("name", "Order")[:31]

            rate = self._get_rate()
            yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

            headers = [
                "№", "Фото", "Название (RU)", "Цена ¥", "Цена $",
                "SKU", "Заказчик", "Описание", "Кол-во",
                "Сумма товар $", "Ссылка", "Статус"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            col_widths = {
                "A": 5, "B": 14, "C": 40, "D": 10, "E": 10,
                "F": 30, "G": 15, "H": 35, "I": 8, "J": 12,
                "K": 50, "L": 10
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            for i, p in enumerate(self._products):
                row = 2 + i
                qty = p.get("quantity", 1)
                price_usd = round(p["price_yuan"] / rate, 2)
                total_usd = round(price_usd * qty, 2)

                ws.cell(row=row, column=1, value=i + 1)

                image_path = p.get("image_path", "")
                if image_path and Path(image_path).exists():
                    try:
                        img = PILImage.open(image_path)
                        img.thumbnail((90, 90))
                        buf = io.BytesIO()
                        img.save(buf, format="JPEG")
                        buf.seek(0)
                        xl_img = XLImage(buf)
                        xl_img.width = 80
                        xl_img.height = 80
                        ws.add_image(xl_img, f"B{row}")
                    except Exception:
                        ws.cell(row=row, column=2, value="Ошибка")
                else:
                    ws.cell(row=row, column=2, value="Нет фото")

                ws.cell(row=row, column=3, value=p.get("title_ru", "") or p.get("title_zh", ""))
                ws.cell(row=row, column=4, value=p["price_yuan"])
                ws.cell(row=row, column=5, value=price_usd)
                ws.cell(row=row, column=6, value=p.get("sku_selected", ""))
                ws.cell(row=row, column=7, value=p.get("customer", ""))

                desc_cell = ws.cell(row=row, column=8, value=p.get("description", ""))
                desc_cell.fill = yellow_fill

                ws.cell(row=row, column=9, value=qty)
                ws.cell(row=row, column=10, value=total_usd)

                link_cell = ws.cell(row=row, column=11, value=p.get("full_url", ""))
                link_cell.hyperlink = p.get("full_url", "")
                link_cell.style = "Hyperlink"

                ws.cell(row=row, column=12, value="Есть" if p.get("status", True) else "Нет")

                ws.row_dimensions[row].height = 70

            wb.save(path)
            self.order_label.setText(f"Товары сохранены: {path}")
        except ImportError:
            self.order_label.setText("Ошибка: openpyxl или Pillow не установлены")
        except Exception as exc:
            self.order_label.setText(f"Ошибка: {exc}")

    def _calc_customer_cost(self, name: str) -> float:
        total = 0.0
        rate = self._get_rate()
        for p in self._products:
            if p.get("customer") == name:
                qty = p.get("quantity", 1)
                total += p["price_yuan"] * qty
        return round(total / rate, 2) if rate > 0 else 0.0

    def _export_weights(self) -> None:
        participants = self._order_data.get("participants", [])
        if not participants:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт веса", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Вес"

            pkg_share = self._packaging_weight_g / 1000.0 / len(participants) if participants else 0.0
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

            headers = ["Участник", "Вес (кг)", "Упаковка (кг)", "Итого вес (кг)", "Доставка ($)", "Сумма товаров ($)", "Итого ($)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            col_widths = {"A": 20, "B": 14, "C": 16, "D": 16, "E": 16, "F": 18, "G": 14}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            total_weight = 0.0
            total_delivery = 0.0
            total_goods = 0.0
            total_all = 0.0

            for i, name in enumerate(participants):
                row = 2 + i
                personal = self._weights.get(name, 0.0)
                total_kg = personal + pkg_share
                delivery = total_kg * self._shipping_rate_usd
                goods = self._calc_customer_cost(name)
                grand = goods + delivery
                total_weight += total_kg
                total_delivery += delivery
                total_goods += goods
                total_all += grand

                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=round(personal, 2))
                ws.cell(row=row, column=3, value=round(pkg_share, 2))
                ws.cell(row=row, column=4, value=round(total_kg, 2))
                ws.cell(row=row, column=5, value=round(delivery, 2))
                ws.cell(row=row, column=6, value=goods)
                ws.cell(row=row, column=7, value=round(grand, 2))

            total_row = 2 + len(participants)
            for col, val in enumerate([
                "Итого",
                round(total_weight, 2),
                round(pkg_share * len(participants), 2),
                round(total_weight, 2),
                round(total_delivery, 2),
                round(total_goods, 2),
                round(total_all, 2),
            ], 1):
                cell = ws.cell(row=total_row, column=col, value=val)
                cell.font = bold_font
                cell = ws.cell(row=total_row, column=col, value=val)
                cell.font = bold_font

            wb.save(path)
            self.order_label.setText(f"Вес сохранён: {path}")
        except ImportError:
            self.order_label.setText("Ошибка: openpyxl не установлен")
        except Exception as exc:
            self.order_label.setText(f"Ошибка: {exc}")
